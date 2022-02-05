from tkinter import *
from pathlib import Path
import os
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter


def click():
    datatext = dataText.get()  # edit text for data path
    referencetext = referenceText.get()  # edit text for ref path
    flag1 = 0
    flag2 = 0

    if datatext == "" or referencetext == "":
        l1 = Label(window, text="Paths are empty.                                          ").grid(row=4, column=1,
                                                                                                   sticky=W)
        return

    if Path(datatext).is_dir() == True:
        flag1 = 1

    else:
        print("Windows cannot access the specified path.")
        l2 = Label(window, text="Windows cannot access the specified path.", bg='red').grid(row=4, column=1, sticky=W)
        return

    if Path(referencetext).is_dir() == True:
        flag2 = 1
    else:
        print("Windows cannot access the specified path.")
        l3 = Label(window, text="Windows cannot access the specified path.", bg='red').grid(row=4, column=1, sticky=W)
        return

    if flag1 == 1 and flag2 == 1:  # if both paths are valid

        dataFiles = os.listdir(datatext)  # list all files in this path in dataFiles variable
        print(dataFiles)
        referenceFiles = os.listdir(referencetext)
        print(referenceFiles)

        count = 0 # num of excel files in path
        count2 = len(dataFiles)
        j = 0
        substring = 'xlsx'
        arr = [] # the indexes of excel files

        # loop to get the excel files index
        while count2 > 0:

            if substring in dataFiles[j]:
                print(dataFiles[j])
                arr.append(j)
                count = count + 1
            count2 = count2 - 1
            j = j + 1





        i = 0
        print(count)

        while count > 0:

            # ###### put file name + path in one var########
            temp = '\\'
            filepath = datatext + temp
            refpath = referencetext + temp

            filepath = filepath + dataFiles[arr[i]]
            refpath = refpath + referenceFiles[arr[i]]

            # ######copy and paste from ref to build#######
            buildworkbook = load_workbook(filepath)
            refworkbook = load_workbook(refpath)
            buildworksheet = buildworkbook.worksheets[1]
            refworksheet = refworkbook.worksheets[0]

            # ####frist matrix#####
            for row in range(2, 7):
                for col in range(1, 3):
                    char = get_column_letter(col)
                    val = refworksheet[char + str(row)].value
                    buildworksheet[char + str(row)] = val

            # ####second matrix#####
            for row2 in range(2, 8):
                for col2 in range(6, 10):
                    char2 = get_column_letter(col2)
                    val2 = refworksheet[char2 + str(row2)].value
                    buildworksheet[char2 + str(row2)] = val2

            count = count - 1
            i = i + 1

            buildworkbook.save(filepath)
            refworkbook.close()
            buildworkbook.close()

        l4 = Label(window, text="Done.                                                                          ",
                   bg='green').grid(row=4, column=1, sticky=W)


# ######   UI   #######
window = Tk()
window.title("DBSS Ref")
window.geometry('600x150')

Label(window, text="DBSS Reference", fg="blue", font="none 12 bold").grid(row=1, column=0, sticky=W)

Label(window, text="Build:").grid(row=2, column=0, sticky=W)
dataText = Entry(window, width=50)
dataText.grid(row=2, column=1, sticky=W)

Label(window, text="Reference:").grid(row=3, column=0, sticky=W)
referenceText = Entry(window, width=50)
referenceText.grid(row=3, column=1, sticky=W)

Button(window, text="Start", width=6, command=click).grid(row=6, column=0, sticky=W)

window.mainloop()
